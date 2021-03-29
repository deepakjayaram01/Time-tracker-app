from tkinter import *
import tkinter as tk
from tkinter import ttk
import time
from datetime import datetime
import sys
import os
import openpyxl,xlrd
from openpyxl import Workbook
from openpyxl import load_workbook
import pathlib
import pandas as pd
import subprocess



#creating root
root = Tk()
root.title("Task Tracker")
root.pack_propagate(True)
root.resizable(0,0)
root.configure(bg="grey35")
root.geometry("600x400")
root.iconbitmap("icon.ico")

style = ttk.Style()
style.theme_use("clam")
  

#creating original file
originalfile = pathlib.Path("Task_backend.xlsx")
if originalfile.exists():
    pass
       
else:
    originalfile = Workbook()
    sheet = originalfile["Sheet"]
    #sheet.title = originalfile["Default_remove"]
    
    #add row headers
    sheet.cell(column=1, row=1, value="Task #")
    sheet.cell(column=2, row=1, value="Start time")
    sheet.cell(column=3, row=1, value="End time")
    sheet.cell(column=4, row=1, value="Duration")
    sheet.cell(column=5, row=1, value="Note")
    sheet.cell(column=6, row=1, value="Task title")

    originalfile.save("Task_backend.xlsx")


#Manage tasks window
def managetaskswd():
    root1 = Tk()
    root1.title("Manage tasks")
    root1.geometry("400x400")
    root1.iconbitmap("icon.ico")
    
    frame = Frame(root1)
    frame.pack()
    
    #create new backend file
    file = pathlib.Path("Task_backend.xlsx")
    if file.exists():
        pass
           
    else:
        file = Workbook()
        file.save("Task_backend.xlsx")
    
    #get sheet titles
    file2 = openpyxl.load_workbook("Task_backend.xlsx")
    sheettitles = file2.sheetnames
    print (sheettitles)
    

    
    #adding listbox in a loop
    listbox = Listbox(frame)
    listbox.pack(pady=10, ipadx=200)
    
    for sheet in sheettitles:
        listbox.insert(END, sheet)
    

    
    #add task 
    def add_task():
        taskentry.get()
        print (taskentry.get())
        listbox.insert(END, taskentry.get())
       
        file1 = load_workbook("Task_backend.xlsx")
        file1.create_sheet(title=taskentry.get(), index=0)
        file1.save("Task_backend.xlsx")
        
        active_sheet = file1[taskentry.get()]
        
        #add row headers
        active_sheet.cell(column=1, row=1, value="Task #")
        active_sheet.cell(column=2, row=1, value="Start time")
        active_sheet.cell(column=3, row=1, value="End time")
        active_sheet.cell(column=4, row=1, value="Duration")
        active_sheet.cell(column=5, row=1, value="Note")
        active_sheet.cell(column=6, row=1, value="Task title")
             
        file1.save("Task_backend.xlsx")
        
        combo.delete("0", END)
        sheetlist = file1.sheetnames
        combo["values"] = sheetlist
     
        
     #delet task
    def delete_task():
        selected_sheet = str(listbox.get(listbox.curselection()))
        print (selected_sheet)
        
        file2 = openpyxl.load_workbook("Task_backend.xlsx")
        std = file2[selected_sheet]
        file2.remove(std)
        file2.save("Task_backend.xlsx")
        
        listbox.delete(ANCHOR)
        
        combo.delete("0", END)
        sheetlist = file2.sheetnames
        combo["values"] = sheetlist
        

        
    #task entry box  
    addlabel = Label(root1, text="Add new task title here and click 'Add task'", fg="grey30", font=('Calibri', 9))
    addlabel.pack(pady=5)
    taskentry = Entry(frame)
    taskentry.pack(ipadx=200)
    
    addbutton = Button(root1, text="Add task", command=add_task, width=100, font=('Calibri', 9,'bold'))
    addbutton.pack()
    rmbutton = Button(root1, text="Delete task", command=delete_task, width=100, font=('Calibri', 9,'bold'))
    rmbutton.pack()
    
        
        
    root1.mainloop()
    
    

#Creating menu
my_menu = Menu(root)
root.config(menu=my_menu)


#command to open database excel
def open_db():
    subprocess.Popen(["Task_backend.xlsx"], shell=True)

#function to open summary window
def report_summary():
    
    f = pathlib.Path("Task_summary.xlsx")
    if f.exists():
        os.remove(f)
        
        df = pd.concat(pd.read_excel("Task_backend.xlsx", sheet_name=None), ignore_index=True)
        df.to_excel(r'Task_summary.xlsx', index = False)
        os.system('start "excel" "Task_summary.xlsx"')


    else:
        df = pd.concat(pd.read_excel("Task_backend.xlsx", sheet_name=None), ignore_index=True)
        df.to_excel(r'Task_summary.xlsx', index = False)
        os.system('start "excel" "Task_summary.xlsx"')


    

#Creating File menu
file_menu = Menu(my_menu,tearoff=0)
my_menu.add_cascade(label="File", menu=file_menu)
file_menu.add_command(label="Open database", command=open_db)
file_menu.add_separator()
file_menu.add_command(label="Exit", command=root.quit)

#Creating Reports menu
reports_menu = Menu(my_menu, tearoff=0)
my_menu.add_cascade(label="Report", menu=reports_menu)
reports_menu.add_command(label="Summary", command=report_summary)

        
#Adding timer and start/pause/reset buttons
global count
count =0

  
def reset():
    global count
    count=1
    t.set('00:00:00')
    
    bt1["state"] = tk.NORMAL
    bt2["state"] = tk.DISABLED
    bt3["state"] = tk.NORMAL
    bt4["state"] = tk.DISABLED
        
def start():
    global count
    count=0
    timer()
    
    global starttime
    global starttime1
    starttime = datetime.now()    
    starttime1 = datetime.strftime(starttime, '%m/%d/%Y %H:%M:%S')
    
    bt1["state"] = tk.DISABLED
    bt2["state"] = tk.NORMAL
    bt3["state"] = tk.NORMAL
 

def pause():
    global count
    count=1
    
    global endtime1
    endtime = datetime.now()
    endtime1 = datetime.strftime(endtime, '%m/%d/%Y %H:%M:%S')
    
    global taskduration
    taskduration = endtime - starttime
    
    
    bt1["state"] = tk.DISABLED
    bt2["state"] = tk.DISABLED
    bt3["state"] = tk.NORMAL
    bt4["state"] = tk.NORMAL

           
def timer():
    global count
    if(count==0):
        d = str(t.get())
        h,m,s = map(int,d.split(":"))
         
        h = int(h)
        m=int(m)
        s= int(s)
        if(s<59):
            s+=1
        elif(s==59):
            s=0
            if(m<59):
                m+=1
            elif(m==59):
                m=0
                h+=1
        if(h<10):
            h = str(0)+str(h)
        else:
            h= str(h)
        if(m<10):
            m = str(0)+str(m)
        else:
            m = str(m)
        if(s<10):
            s=str(0)+str(s)
        else:
            s=str(s)
        d=h+":"+m+":"+s
         
         
        t.set(d)
        if(count==0):
            frame1.after(1000,timer)
             
 
 
     
#creating top frame

frame1 = LabelFrame(root, bd=1, bg="grey64", padx=5, pady=5)
frame1.pack(fill="both")
frame1.pack(padx=5, pady=5)  
#Add task button
addtskbutton = Button(frame1, text="Manage tasks", command=managetaskswd, font=('Calibri', 9,'bold'))
addtskbutton.grid(column=6, row=0, sticky=E, padx=5)




def make_entry():
    df = pd.read_excel("Task_backend.xlsx", sheet_name=combo_select)
    count_row = df.shape[0]
    print (count_row)
    new_row_count = count_row + 1
    
    
    file = openpyxl.load_workbook("Task_backend.xlsx")
    sheet=file[combo_select]
    

    sheet.cell(column=1, row=sheet.max_row+1, value=new_row_count)
    sheet.cell(column=2, row=sheet.max_row, value=starttime1)
    sheet.cell(column=3, row=sheet.max_row, value=endtime1)
    sheet.cell(column=4, row=sheet.max_row, value=taskduration)
    sheet.cell(column=5, row=sheet.max_row, value=noteentry.get())
    sheet.cell(column=6, row=sheet.max_row, value=combo.get())
    file.save("Task_backend.xlsx")
    
    tv1.delete(*tv1.get_children())
    
  
    df = pd.read_excel("Task_backend.xlsx", sheet_name=combo_select)
    tv1["column"] = list(df.columns)
    tv1.column("#0", width="50")
    tv1.column("#1", width="50")
    tv1.column("#2", width="130")
    tv1.column("#3", width="130")
    tv1.column("#4", width="130")
    tv1.column("#5", width="300")
    tv1["show"] = "headings"
    #loop through column list in headers
    for column in tv1["column"]:
        tv1.heading(column, text=column)
        
    #adding data to treeview
    df_rows = df.to_numpy().tolist()
    for row in df_rows:
        tv1.insert("", "end", values=row)
        
    #making paramenters default
    noteentry.delete("0", END)
    t.set("00:00:00")
    bt4["state"] = tk.DISABLED
    bt1["state"] = tk.NORMAL
    
    
        

#Notes entry
notelabel = Label(frame1, text="Notes", fg="black", bg="grey64")
notelabel.grid(column=4, row=2, sticky=E)
noteentry = Entry(frame1, width=59)
noteentry.grid(column=5, row=2, sticky=W, ipady=3)


#Make entry button
bt4 = Button(frame1,text="Make entry", padx=8, command=make_entry, state=DISABLED, font=('Calibri', 9,'bold'))
bt4.grid(column=6, row=2, sticky=E, padx=5)




#adding start-pause-reset buttons and timer
t = StringVar()
t.set("00:00:00")

img1 = tk.PhotoImage(file=r"playbutton.png")
img2 = tk.PhotoImage(file=r"Pausebutton.png")
img3 = tk.PhotoImage(file=r"Refreshbutton.png")
img1=img1.subsample(5)
img2=img2.subsample(5)
img3=img3.subsample(5)
lb = Label(frame1,textvariable=t, fg="black", relief="sunken", borderwidth=0, bg="grey64", font=("Sans 18"))
bt1 = Button(frame1, image=img1, width=25, height=25, bg="grey83", command=start)
bt2 = Button(frame1, image=img2, width=25, height=25, bg="grey83", command=pause, state=DISABLED)
bt3 = Button(frame1, image=img3, width=25, height=25, bg="grey83", command=reset, state=DISABLED)

lb.grid(column=1, row=0, columnspan=3)
bt1.grid(column=1, row=2, sticky=W)
bt2.grid(column=2, row=2)
bt3.grid(column=3, row=2)




#Creating data table
frame2 = LabelFrame(root, bd=1, bg="grey60", padx=5, pady=5)
frame2.pack(fill="both", expand="yes")
frame2.pack(padx=5, pady=5)            



#styling the treeview
style = ttk.Style()
style.configure("mystyle.Treeview", highlightthickness=0, bd=0, font=('ariel', 9)) # Modify the font of the body
style.configure("mystyle.Treeview.Heading", font=('Calibri', 8,'bold')) # Modify the font of the headings

#Treeview
tv1 = ttk.Treeview(frame2,style="mystyle.Treeview")

tv1.place(relheight=1, relwidth=1)
treescrolly = Scrollbar(frame2, orient="vertical", command=tv1.yview)  
treescrollx = Scrollbar(frame2, orient="horizontal", command=tv1.xview) 
tv1.configure(xscrollcommand=treescrollx.set, yscrollcommand=treescrolly.set)   
treescrollx.pack(side="bottom", fill="x")  
treescrolly.pack(side="right", fill="y")


def comboclick(event):
    tv1.delete(*tv1.get_children())
    
    #fetching data from excel
    global combo_select
    combo_select = combo.get()
 
    df = pd.read_excel("Task_backend.xlsx", sheet_name=combo_select)
    count_row = df.shape[0]
  
    
    tv1["column"] = list(df.columns)
    tv1.column("#0", width="50")
    tv1.column("#1", width="50")
    tv1.column("#2", width="130")
    tv1.column("#3", width="130")
    tv1.column("#4", width="130")
    tv1.column("#5", width="300")
    tv1["show"] = "headings"
    #loop through column list in headers
    for column in tv1["column"]:
        tv1.heading(column, text=column)
        
    #adding data to treeview
    df_rows = df.to_numpy().tolist()
    for row in df_rows:
        tv1.insert("", "end", values=row)
    


#creating drop downlist of tasks
wb = openpyxl.load_workbook("Task_backend.xlsx")
sheetlist = wb.sheetnames
print(sheetlist)


Tasklist = Label(frame1, text="Task", fg="black", bg="grey64")
Tasklist.grid(column=4, row=0, sticky=E)
combo = ttk.Combobox(frame1, value=sheetlist, width=57)
combo.current()
combo.bind("<<ComboboxSelected>>", comboclick)
combo.grid(column=5, row=0, sticky=W, ipady=3)


comboselect = combo.get()















                 




root.mainloop()



